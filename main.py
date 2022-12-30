import random
import numpy as np
from openpyxl import Workbook

from constraints import *

NDAYS = 5
NHOURS = 6

ALL_DAYS = list(range(5))
ALL_HOURS = list(range(6))

DAYS = {0: 'LUNEDI', 1: 'MARTEDI', 2: 'MERCOLEDI', 3: 'GIOVEDI', 4: 'VENERDI'}

HOURS = {
    0: '8.30-10.00',
    1: '10.00-11.30',
    2: '11.30-13.00',
    3: '14.30-16.00',
    4: '16.00-17.30',
    5: '17.30-19.00'
}

SUBJECTS = {
    -1: '',
    0: 'ITA',
    1: 'MAT',
    2: 'ING',
    3: 'MUS',
    4: 'EDF',
    5: 'STO',
    6: 'GEO',
    7: 'SCZ',
    8: 'INF',
    9: 'VID'
}

ITA = 0
MAT = 1
ING = 2
MUS = 3
EDF = 4
STO = 5
GEO = 6
SCZ = 7
INF = 8
VID = 9

PRIMA = 1
SECONDA = 2
TERZA = 3
QUARTA = 4
QUINTA = 5
ARANCIONI = 6
ROSSI = 7
VERDI = 8

LUNEDI = 0
MARTEDI = 1
MERCOLEDI = 2
GIOVEDI = 3
VENERDI = 4

MORNING = [0, 1, 2]
AFTERNOON = [3, 4, 5]
SCHOOL_TIME = [0, 1, 2, 3]
NOT_TOO_LATE = [0, 1, 2, 3, 4]
LATE = [5]
NOT_TOO_EARLY = [1,2,3,4]

MATERIE_PRIMA = [ITA, ITA, ITA, ITA, MAT, MAT, MAT, ING, MUS, EDF]
MATERIE_SECONDA = MATERIE_PRIMA
MATERIE_TERZA = [
    ITA, ITA, ITA, ITA, ITA, ITA, MAT, MAT, MAT, MAT, ING, ING, EDF, MUS, INF
]
MATERIE_QUARTA = [ING, ING, EDF, MUS, INF]
MATERIE_QUINTA = MATERIE_TERZA
MATERIE_INFANZIA = [VID]

# giorno -> orario -> classe


class Classe:

    def __init__(self, subjects, timetable=None):
      if timetable is None:
        timetable = np.array([[-1 for _ in ALL_HOURS] for _ in ALL_DAYS])

      self.timetable = timetable
      self.subjects = subjects

    def is_full(self):
      return len(self.subjects) == 0

    def is_free(self, d, h):
    	return self.timetable[d, h] == -1

    def schedule_next_lesson(self, d, h):
    	s = self.subjects.pop(0)
    	self.timetable[d, h] = s

    def duplicate(self):
      return Classe(self.subjects.copy(), 
                    timetable=self.timetable.copy())

    def schedule_lesson_at_random_time(self):
      classe = self.duplicate()

      found = False

      while not found:

        d, h = random.choice(ALL_DAYS), random.choice(ALL_HOURS)
        found = classe.is_free(d, h)

        if found:
        	classe.schedule_next_lesson(d, h)

      return classe

    def print_timetable(self):
        printable = self.timetable.copy().astype(np.chararray)
        for id, subject in SUBJECTS.items():
            printable[printable == id] = subject

        print(printable)

    def subject_to_string(self, subject):
      return SUBJECTS[subject]

    def print_on_excel(self, sheet, index):

      for id, hour in HOURS.items():
        sheet.cell(1 + index + id, 1).value = hour

      for id, day in DAYS.items():
        sheet.cell(index, 2 + id).value = day

      for row in ALL_HOURS:
        for col in ALL_DAYS:
          sheet.cell(1 + index + row, 2 + col).value = self.subject_to_string(self.timetable[col, row])
      



classes = {
        PRIMA: Classe(MATERIE_PRIMA),
        SECONDA: Classe(MATERIE_SECONDA),
        TERZA: Classe(MATERIE_TERZA),
        QUARTA: Classe(MATERIE_QUARTA),
        QUINTA: Classe(MATERIE_QUINTA),
        ARANCIONI: Classe(MATERIE_INFANZIA),
        VERDI: Classe(MATERIE_INFANZIA),
        ROSSI: Classe(MATERIE_INFANZIA)

    }



constraints = [
	NoMoreThanNHoursForTeacher(2, [ITA, MAT], QUINTA),
	NoMoreThanNHoursForTeacher(2, [ITA, MAT], TERZA),
	NoSameTime(ROSSI, SECONDA),
	NoSameTime(ROSSI, TERZA),
	NoSameTime(ROSSI, QUARTA),
	NoSameTime(ROSSI, QUINTA),
	NoSameTime(ARANCIONI, PRIMA),
	NoSameTime(VERDI, SECONDA),
	NoMoreThanNClassesADay(3),
	NoMoreThanNClassesADayOfSubject(2, ITA),
	NoMoreThanNClassesADayOfSubject(2, MAT),
	NoMoreThanNClassesADayOfSubject(1, ITA, [PRIMA]),
    NoMoreThanNClassesADayOfSubject(1, MAT, [SECONDA]),
    NoMoreThanNClassesADayOfSubject(1, MAT, [PRIMA]),
    # FACOLTATIVI
    NoMoreThanNClassesADay(1, [QUARTA]),
    SubjectInCertainDaysAndHours(MAT, ALL_DAYS, NOT_TOO_LATE, classes = [PRIMA]),
    NoSameSubjectForAnyone(ING),
    NoSameSubjectForAnyone(MUS),
    NoSameSubjectForAnyone(EDF),
    NoSameSubjectForAnyone(INF),
    SubjectInCertainDaysAndHours(MUS, [MERCOLEDI, GIOVEDI, VENERDI],
                                 NOT_TOO_LATE),
    SubjectInCertainDaysAndHours(INF, [LUNEDI, MARTEDI, GIOVEDI],
                                 NOT_TOO_EARLY),
    SubjectInCertainDaysAndHours(ING, ALL_DAYS, MORNING),
    NoSameTeacherForTwoClasses(PRIMA, ITA, SECONDA, MAT),
    SubjectInCertainDaysAndHours(EDF, ALL_DAYS, MORNING),
    SubjectInCertainDaysAndHours(MUS, ALL_DAYS, MORNING, classes = [QUARTA]),
    SubjectInCertainDaysAndHours(INF, ALL_DAYS, MORNING, classes = [QUARTA]),
    SubjectInCertainDaysAndHours(ITA, ALL_DAYS, SCHOOL_TIME, classes = [TERZA]),
    SubjectInCertainDaysAndHours(MAT, ALL_DAYS, SCHOOL_TIME, classes = [TERZA]),
    NotBusyOnDaysAndHours(PRIMA, ALL_DAYS, LATE),
    NotBusyOnDaysAndHours(SECONDA, ALL_DAYS, LATE),
    SubjectInCertainDaysAndHours(ITA, ALL_DAYS, MORNING, classes = [QUINTA]),
    SubjectInCertainDaysAndHours(MAT, ALL_DAYS, MORNING, classes = [QUINTA]),
    SubjectInCertainDaysAndHours(ITA, [LUNEDI, MARTEDI, MERCOLEDI, VENERDI], MORNING, classes = [PRIMA]),
    SubjectInCertainDaysAndHours(MAT, ALL_DAYS, MORNING, classes = [PRIMA]),
    SubjectInCertainDaysAndHours(MAT, [MARTEDI, GIOVEDI, VENERDI], MORNING, classes = [SECONDA]),
    SubjectInCertainDaysAndHours(ITA, ALL_DAYS, SCHOOL_TIME, classes = [SECONDA]),
    SubjectInCertainDaysAndHours(VID, ALL_DAYS, MORNING, classes = [ARANCIONI, VERDI, ROSSI])

]




class Calendar:
  def __init__(self, classes):
    self.classes = classes

  def is_full(self):
    return np.all([ classe.is_full() for classe in self.classes.values() ])

  def schedule_random_lesson(self):
    classes = self.classes.copy()
    not_full = [ (name, classe) for name, classe in classes.items() if not classe.is_full() ]

    name, classe = random.choice(not_full)
    classes[name] = classe.schedule_lesson_at_random_time()
    return Calendar(classes)

  def satisfies(self, constraints):
  	return np.all([ c(self.classes) for c in constraints ])

  def save_as_excel(self):
    wb = Workbook()
    sheet = wb.active

    for id, classe in self.classes.items():
     classe.print_on_excel(sheet, (id - 1) * 9 + 1)

    # list(self.classes.values())[0].print_on_excel(sheet, 1)
    sheet.column_dimensions['A'].width = 15
    for column in range(1, 10):
    	column_char = str(chr(64+column))
    	sheet.column_dimensions[column_char].width = 10
    wb.save(filename="orario.xlsx")

  def save_as_numpy(self):
    for id, classe in self.classes.items():
       np.save('classi/classe' + str(id) + '.npy', classe.timetable)


class CalendarSolver:
	def __init__(self, iteration_number = 0, iteration_limit = 1000):
		self.iteration_number = iteration_number
		self.iteration_limit = iteration_limit


	def generate_empty_calendar(self):
		return Calendar(classes)

	def next_step(self, calendar):
		if self.iteration_number > self.iteration_limit:
			self.iteration_number = 0
			print('Starting over...')
			return self.generate_empty_calendar()

		self.iteration_number += 1
		return calendar.schedule_random_lesson()

     


cs = CalendarSolver()
calendar = cs.generate_empty_calendar()

while not calendar.is_full():

	modified_calendar = cs.next_step(calendar)

	if modified_calendar.satisfies(constraints):

		calendar = modified_calendar



calendar.save_as_excel()
calendar.save_as_numpy()

for c in calendar.classes.values():
    c.print_timetable()



# calendar = load_calendar()
# calendar.classes[TERZA].timetable[]